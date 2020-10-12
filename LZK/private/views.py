import logging
from itertools import islice

from braces.views import SuperuserRequiredMixin
from crispy_forms.bootstrap import FormActions, StrictButton
from crispy_forms.helper import FormHelper
from crispy_forms.layout import ButtonHolder, Fieldset, Layout, Submit
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.mail import EmailMultiAlternatives
from django.http import HttpResponseNotAllowed, HttpResponseRedirect
from django.template import Context
from django.template.loader import get_template
from django.urls import reverse_lazy as reverse
from django.utils.translation import gettext_lazy as _
from django.views.generic import (
    CreateView,
    DetailView,
    FormView,
    ListView,
    TemplateView,
    View,
)
from django.views.generic.detail import SingleObjectMixin
from django_filters.views import FilterView
from django_tables2.views import SingleTableMixin
from openpyxl import load_workbook
from psqlextra.query import ConflictAction
from psqlextra.util import postgres_manager

from .. import models
from ..mixins import FilterFormHelperMixin
from . import filters, forms, tables
from .conf import settings

logger = logging.getLogger(__name__)


class IndexView(LoginRequiredMixin, TemplateView):
    template_name = "LZK/private/index.html"


class ImportView(LoginRequiredMixin, SuperuserRequiredMixin, FormView):
    template_name = "LZK/private/import.html"
    form_class = forms.ImportForm
    extra_context = {
        "sheets": [
            settings.LZK_IMPORT_SHEET_ACRONYMS,
            settings.LZK_IMPORT_SHEET_UFIDS,
            settings.LZK_IMPORT_SHEET_OBJECTIVES,
        ]
    }
    success_url = reverse("private:index")

    def form_valid(self, form):
        wb = load_workbook(
            form.cleaned_data.get("file"),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        acronyms = {
            k.value.upper(): v.value
            for v, k, *_ in wb.get_sheet_by_name(settings.LZK_IMPORT_SHEET_ACRONYMS)
        }
        ufid_sheet = wb.get_sheet_by_name(settings.LZK_IMPORT_SHEET_UFIDS)
        ufid_data = (
            {"id": x[2].value, "name": x[3].value}
            for x in islice(ufid_sheet.iter_rows(), 1, None)
        )
        models.UFID.objects.on_conflict(["id"], ConflictAction.UPDATE).bulk_insert(
            ufid_data
        )

        objective_sheet = wb.get_sheet_by_name(settings.LZK_IMPORT_SHEET_OBJECTIVES)
        ability_data = list()
        level_data = dict()
        activity_data = dict()
        cl_data = dict()
        skill_data = list()
        skill_activity_map = dict()
        subject_data = dict()
        system_data = dict()
        study_field_data = dict()
        ability_level_data = list()
        ability_subject_data = list()
        ability_system_data = list()
        ability_ufid_data = list()
        ability_study_field_data = dict()
        symptom_data = list()
        symptom_subject_data = list()
        for row in islice(objective_sheet.iter_rows(), 1, None):
            if row[10].value.strip().lower() == settings.LZK_IMPORT_VALUE_TRUE.lower():
                # Symptom
                symptom_data.append(
                    {
                        "pk": row[0].value,
                        "name": row[1].value.strip(),
                        "public": row[14].value.strip().lower
                        == settings.LZK_IMPORT_VALUE_TRUE.lower(),
                    }
                )
                # Subjects
                if row[7].value:
                    for s in filter(
                        bool, map(lambda s: s.strip(), row[7].value.split(","))
                    ):
                        if s.upper() not in subject_data:
                            subject_data[s.upper()] = {
                                "name": acronyms.get(s.upper(), s)
                            }
                        symptom_subject_data.append(
                            {"symptom_id": row[0].value, "subject_id": s.upper()}
                        )
            elif bool(row[11].value):
                # Competence levels
                if row[3].value and row[6].value:
                    cl = row[3].value.strip()
                    if cl.upper() not in cl_data:
                        cl_data[cl.upper()] = {
                            "name": acronyms.get(cl.upper(), cl),
                        }
                    act = row[6].value.strip()
                    if act not in activity_data:
                        activity_data[act] = cl.upper()
                    skill_data.append(
                        {"pk": row[0].value, "name": row[1].value.strip(),}
                    )
                    skill_activity_map[row[0].value] = act
            else:
                ability_data.append(
                    {
                        "pk": row[0].value,
                        "name": row[1].value.strip(),
                        "depth": row[2].value,
                        "subject_related": row[9].value.strip().lower()
                        == settings.LZK_IMPORT_VALUE_TRUE.lower(),
                        "public": row[14].value.strip().lower()
                        == settings.LZK_IMPORT_VALUE_TRUE.lower(),
                    }
                )
                # Levels
                if row[3].value:
                    for l in map(lambda l: l.strip(), row[3].value.split(",")):
                        if l.upper() not in level_data:
                            level_data[l.upper()] = {"name": acronyms.get(l.upper(), l)}
                        ability_level_data.append(
                            {"ability_id": row[0].value, "level_id": l.upper()}
                        )
                # Subjects
                if row[7].value:
                    for s in filter(
                        bool, map(lambda s: s.strip(), row[7].value.split(","))
                    ):
                        if s.upper() not in subject_data:
                            subject_data[s.upper()] = {
                                "name": acronyms.get(s.upper(), s)
                            }
                        ability_subject_data.append(
                            {"ability_id": row[0].value, "subject_id": s.upper()}
                        )
                # Systems
                if row[12].value:
                    for s in map(lambda s: s.strip(), row[12].value.split(",")):
                        if s.upper() not in system_data:
                            system_data[s.upper()] = {
                                "name": acronyms.get(s.upper(), s)
                            }
                        ability_system_data.append(
                            {"ability_id": row[0].value, "system_id": s.upper()}
                        )
                # UFIDs
                if row[15].value:
                    if isinstance(row[15].value, int):
                        values = [row[15].value]
                    elif isinstance(row[15].value, float):
                        values = map(
                            lambda u: int(u.strip()),
                            filter(bool, str(row[15].value).split(".")),
                        )
                    else:
                        values = map(
                            lambda u: int(u.strip()),
                            filter(bool, row[15].value.split(",")),
                        )
                    for u in values:
                        ability_ufid_data.append(
                            {"ability_id": row[0].value, "ufid_id": u}
                        )
                # Study field
                if row[16].value:
                    sf = row[16].value.strip()
                    if sf.upper() not in study_field_data:
                        study_field_data[sf.upper()] = {
                            "name": acronyms.get(sf.upper(), sf)
                        }
                    ability_study_field_data[row[0].value] = sf.upper()

        models.StudyField.objects.on_conflict(
            ["pk"], ConflictAction.UPDATE
        ).bulk_insert(({**v, **{"pk": k}} for k, v in study_field_data.items()))

        ability_data_update = [
            {**o, **{"study_field_id": ability_study_field_data.get(o.get("pk")),},}
            for o in ability_data
            if o.get("pk") in ability_study_field_data
        ]

        for k, v in cl_data.items():
            models.CompetenceLevel.objects.get_or_create(pk=k, defaults=v)

        models.Activity.objects.on_conflict(
            ["name"], ConflictAction.UPDATE
        ).bulk_insert(
            ({"name": k, "competence_level_id": v} for k, v in activity_data.items())
        )
        activities_pk = {a.name: a.pk for a in models.Activity.objects.all()}

        skill_data_update = [
            {
                **s,
                **{
                    "activity_id": activities_pk.get(
                        skill_activity_map.get(s.get("pk"))
                    ),
                },
            }
            for s in skill_data
            if s.get("pk") in skill_activity_map
        ]

        models.Skill.objects.on_conflict(["pk"], ConflictAction.UPDATE).bulk_insert(
            skill_data_update
        )

        objs = models.Ability.objects.on_conflict(
            ["pk"], ConflictAction.UPDATE
        ).bulk_insert(ability_data_update)
        # models.Ability.subjects.through.objects.filter(
        #    ability__pk__in=pks
        # ).delete()

        syms = models.Symptom.objects.on_conflict(
            ["pk"], ConflictAction.UPDATE
        ).bulk_insert(symptom_data)
        # models.Symptom.subjects.through.objects.filter(
        #    symptom__pk__in=pks
        # ).delete()

        models.Subject.objects.on_conflict(["pk"], ConflictAction.UPDATE).bulk_insert(
            ({**v, **{"pk": k}} for k, v in subject_data.items())
        )

        with postgres_manager(models.Ability.subjects.through) as manager:
            manager.get_queryset().bulk_insert(ability_subject_data)

        with postgres_manager(models.Symptom.subjects.through) as manager:
            manager.get_queryset().bulk_insert(symptom_subject_data)

        models.Level.objects.on_conflict(["pk"], ConflictAction.UPDATE).bulk_insert(
            ({**v, **{"pk": k}} for k, v in level_data.items())
        )
        models.System.objects.on_conflict(["pk"], ConflictAction.UPDATE).bulk_insert(
            ({**v, **{"pk": k}} for k, v in system_data.items())
        )

        # models.Ability.levels.through.objects.filter(
        #    ability__pk__in=pks
        # ).delete()
        with postgres_manager(models.Ability.levels.through) as manager:
            manager.get_queryset().bulk_insert(ability_level_data)

        # models.Ability.systems.through.objects.filter(
        #    ability__pk__in=pks
        # ).delete()
        with postgres_manager(models.Ability.systems.through) as manager:
            manager.get_queryset().bulk_insert(ability_system_data)

        # models.Ability.ufids.through.objects.filter(
        #    ability__pk__in=pks
        # ).delete()
        with postgres_manager(models.Ability.ufids.through) as manager:
            manager.get_queryset().bulk_insert(ability_ufid_data)

        return super().form_valid(form)


class ListAbilityView(LoginRequiredMixin, ListView):
    model = models.Ability
    paginate_by = 50
    template_name = "LZK/private/ability/list.html"


class ListFeedbackView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = models.Feedback
    template_name = "LZK/private/feedback/list.html"
    table_class = tables.FeedbackTable
    filterset_class = filters.FeedbackFilter

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.filter(active=True)


class CreateFeedbackView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = models.Feedback
    form_class = forms.FeedbackForm
    template_name = "LZK/private/feedback/create.html"
    success_url = reverse("private:feedback-list")
    permission_required = "LZK.add_feedback"

    def form_valid(self, form):
        response = super().form_valid(form)
        plaintext = get_template("LZK/private/email/feedback.txt")
        html = get_template("LZK/private/email/feedback.html")
        for contact in self.object.university.contact_set.all():
            logger.info(f"Sending mail: {self.object} to {contact}")
            d = {
                "contact": contact,
                "object": self.object,
                "request": self.request,
            }
            msg = EmailMultiAlternatives(
                _("LZK: Request for feedback"),
                plaintext.render(d),
                settings.LZK_EMAIL_FROM,
                [contact.email],
            )
            msg.attach_alternative(html.render(d), "text/html")
            msg.send()
        return response


class DetailFeedbackView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = models.Feedback
    template_name = "LZK/private/feedback/detail.html"
    permission_required = "LZK.view_feedback"


class CloseFeedbackView(
    LoginRequiredMixin, PermissionRequiredMixin, SingleObjectMixin, View
):
    model = models.Feedback
    permission_required = "LZK.change_feedback"

    def post(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.active = False
        obj.save()
        url = reverse("private:feedback-list")
        return HttpResponseRedirect(url)


class AbilityCommentView(
    LoginRequiredMixin,
    PermissionRequiredMixin,
    SingleTableMixin,
    SingleObjectMixin,
    FilterView,
):
    model = models.Feedback
    template_name = "LZK/private/feedback/abilities.html"
    table_class = tables.AbilityCommentTable
    permission_required = "LZK.view_feedback"
    filterset_class = filters.AbilityCommentFilter

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().get(request, *args, **kwargs)

    def get_filterset_kwargs(self, filterset_class):
        kwargs = super().get_filterset_kwargs(filterset_class)
        kwargs["queryset"] = self.object.abilitycomment_set.filter(
            status=self.table_class.Meta.model.OPEN
        )
        return kwargs


class ChangeAbilityCommentView(
    LoginRequiredMixin, PermissionRequiredMixin, SingleObjectMixin, View
):
    model = models.AbilityComment
    permission_required = "LZK.change_feedback"

    states = {
        "accept": model.ACCEPTED,
        "discard": model.DISCARDED,
    }

    def post(self, request, action, *args, **kwargs):
        state = self.states.get(action)
        if not state:
            return HttpResponseNotAllowed()
        obj = self.get_object()
        obj.status = state
        obj.save()
        url = reverse(
            "private:feedback-detail-abilities", kwargs={"pk": obj.feedback.pk}
        )
        return HttpResponseRedirect(url)


class SymptomCommentView(
    LoginRequiredMixin,
    PermissionRequiredMixin,
    SingleTableMixin,
    SingleObjectMixin,
    FilterView,
):
    model = models.Feedback
    template_name = "LZK/private/feedback/symptoms.html"
    table_class = tables.SymptomCommentTable
    permission_required = "LZK.view_feedback"
    filterset_class = filters.SymptomCommentFilter

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().get(request, *args, **kwargs)

    def get_filterset_kwargs(self, filterset_class):
        kwargs = super().get_filterset_kwargs(filterset_class)
        kwargs["queryset"] = self.object.symptomcomment_set.filter(
            status=self.table_class.Meta.model.OPEN
        )
        return kwargs


class ChangeSymptomCommentView(
    LoginRequiredMixin, PermissionRequiredMixin, SingleObjectMixin, View
):
    model = models.SymptomComment
    permission_required = "LZK.change_feedback"

    states = {
        "accept": model.ACCEPTED,
        "discard": model.DISCARDED,
    }

    def post(self, request, action, *args, **kwargs):
        state = self.states.get(action)
        if not state:
            return HttpResponseNotAllowed()
        obj = self.get_object()
        obj.status = state
        obj.save()
        url = reverse(
            "private:feedback-detail-symptoms", kwargs={"pk": obj.feedback.pk}
        )
        return HttpResponseRedirect(url)


class SkillCommentView(
    LoginRequiredMixin,
    PermissionRequiredMixin,
    SingleTableMixin,
    SingleObjectMixin,
    FilterView,
):
    model = models.Feedback
    template_name = "LZK/private/feedback/skills.html"
    table_class = tables.SkillCommentTable
    permission_required = "LZK.view_feedback"
    filterset_class = filters.SkillCommentFilter

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().get(request, *args, **kwargs)

    def get_filterset_kwargs(self, filterset_class):
        kwargs = super().get_filterset_kwargs(filterset_class)
        kwargs["queryset"] = self.object.skillcomment_set.filter(
            status=self.table_class.Meta.model.OPEN
        )
        return kwargs


class ChangeSkillCommentView(
    LoginRequiredMixin, PermissionRequiredMixin, SingleObjectMixin, View
):
    model = models.SkillComment
    permission_required = "LZK.change_feedback"

    states = {
        "accept": model.ACCEPTED,
        "discard": model.DISCARDED,
    }

    def post(self, request, action, *args, **kwargs):
        state = self.states.get(action)
        if not state:
            return HttpResponseNotAllowed()
        obj = self.get_object()
        obj.status = state
        obj.save()
        url = reverse("private:feedback-detail-skills", kwargs={"pk": obj.feedback.pk})
        return HttpResponseRedirect(url)
